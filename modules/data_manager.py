# path: modules/data_manager.py
import os
import pandas as pd
from datetime import datetime, date, time
from typing import Any, Dict, Iterable, Tuple

# ---- yardımcı formatlayıcılar ----
def _fmt_date(d: Any) -> str:
    if isinstance(d, (datetime, date)):
        return d.strftime("%Y-%m-%d")
    return str(d)

def _fmt_time(t: Any) -> str:
    if isinstance(t, time):
        return t.strftime("%H:%M:%S")
    return str(t)

# ---- 1) Ham tablo export ----
def export_current_session_to_excel(
    etud_info: Tuple[str, str, Any, str, time, time, int, int, float, float],
    error_data: Iterable[Dict[str, Any]],
    export_dir: str = "excel_raporlar",
) -> str:
    """
    Hızlı ham tablo çıktısı (satır bazlı). 'Duruş' isimleriyle uyumludur.
    """
    os.makedirs(export_dir, exist_ok=True)

    # unpack
    operator, machine, etud_date, vardiya, start_time, end_time, initial_count, final_count, unit_time, break_time = etud_info
    produced_beds = max(0, (final_count or 0) - (initial_count or 0))

    total_minutes = (
        datetime.combine(date.today(), end_time)
        - datetime.combine(date.today(), start_time)
    ).total_seconds() / 60
    production_time = round(total_minutes - float(break_time or 0), 2)
    planned_production = round(production_time / float(unit_time), 2) if float(unit_time) > 0 else 0

    rows = []
    error_data = list(error_data or [])

    if error_data:
        for err in error_data:
            # geri uyum: "Hata Türü" varsa onu kullan
            durus_turu = err.get("Duruş Türü", err.get("Hata Türü", ""))
            aciklama = err.get("Açıklama", "")
            sure_sn = err.get("Süre (sn)", 0) or 0
            rows.append({
                "Tarih": _fmt_date(etud_date),
                "Makine": machine,
                "Operatör": operator,
                "Vardiya": vardiya,
                "Etüt Başlangıç": _fmt_time(start_time),
                "Etüt Bitiş": _fmt_time(end_time),
                "Başlangıç Sayısı": initial_count,
                "Bitiş Sayısı": final_count,
                "Üretim Süresi (dk)": production_time,
                "Mola (dk)": break_time,
                "Planlanan Üretim": planned_production,
                "Gerçekleşen Üretim": produced_beds,
                "Duruş Türü": durus_turu,
                "Duruş Açıklaması": aciklama,
                "Süre (sn)": sure_sn,
                "Süre (dk)": round(float(sure_sn) / 60, 2),
            })
    else:
        # duruş yoksa tek özet satırı yaz
        rows.append({
            "Tarih": _fmt_date(etud_date),
            "Makine": machine,
            "Operatör": operator,
            "Vardiya": vardiya,
            "Etüt Başlangıç": _fmt_time(start_time),
            "Etüt Bitiş": _fmt_time(end_time),
            "Başlangıç Sayısı": initial_count,
            "Bitiş Sayısı": final_count,
            "Üretim Süresi (dk)": production_time,
            "Mola (dk)": break_time,
            "Planlanan Üretim": planned_production,
            "Gerçekleşen Üretim": produced_beds,
            "Duruş Türü": "",
            "Duruş Açıklaması": "",
            "Süre (sn)": "",
            "Süre (dk)": "",
        })

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(export_dir, f"etut_raporu_{timestamp}.xlsx")
    pd.DataFrame(rows).to_excel(file_path, index=False)
    return file_path


# ---- 2) Tek sayfa “rapor görünümü” export ----
def export_pretty_report(
    etud_info: Tuple[str, str, Any, str, time, time, int, int, float, float],
    error_data: Iterable[Dict[str, Any]],
    summary_df: pd.DataFrame,
    out_path: str,
) -> str:
    """
    Tek sayfa şık rapor (openpyxl ile). 'Duruş' isimleriyle uyumludur.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise ImportError("openpyxl gerekli: pip install openpyxl") from e

    # unpack
    operator, machine, etud_date, vardiya, start_time, end_time, initial_count, final_count, unit_time, break_time = etud_info

    # summary alanları
    s = summary_df.iloc[0].to_dict() if isinstance(summary_df, pd.DataFrame) and not summary_df.empty else {}
    etud_minutes    = s.get("Etüt Süresi (dk)")
    planned_sec     = s.get("Toplam Planlı Süre (sn)") or 0
    unplanned_sec   = s.get("Toplam Plansız Süre (sn)") or 0
    faulty_sec      = s.get("Toplam Hatalı Süre (sn)") or 0
    utilization_pct = s.get("Kapasite Kullanımı (%)")
    actual_qty      = s.get("Gerçekleşen Üretim (adet)")
    planned_qty     = s.get("Planlanan Üretim (adet)")
    performance_pct = s.get("Gerçekleşme Oranı (%)")

    def _num(x, nd=2):
        if x in (None, ""):
            return ""
        try:
            return round(float(x), nd)
        except Exception:
            return x

    wb = Workbook()
    ws = wb.active
    ws.title = "Etüt Raporu"

    # sütun genişlikleri
    for i, w in enumerate([22, 18, 22, 18, 24, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # stiller
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # başlık
    ws.merge_cells("A1:G1")
    ws["A1"] = "Zaman Etüdü Raporu"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = center

    # bilgi bloğu
    rows_info = [
        ("Operatör Adı", operator, "Makine Adı / No", machine, "Etüt Tarihi", _fmt_date(etud_date), ""),
        ("Vardiya", vardiya, "Etüt Başlangıç", _fmt_time(start_time), "Etüt Bitiş", _fmt_time(end_time), ""),
        ("Etüt Öncesi Yatak Sayısı", initial_count, "Etüt Sonrası Yatak Sayısı", final_count, "Toplam Mola (dk)", _num(break_time, 2), ""),
        ("Bir Yatak Süresi (dk)", _num(unit_time, 2), "", "", "", "", ""),
    ]
    start_row = 3
    for r, row in enumerate(rows_info, start=start_row):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = left
        for c in (1, 3, 5):
            ws.cell(row=r, column=c).font = bold

    # özet başlığı
    sum_header_row = start_row + len(rows_info) + 1
    ws.merge_cells(start_row=sum_header_row, start_column=1, end_row=sum_header_row, end_column=7)
    ws.cell(row=sum_header_row, column=1, value="Özet Göstergeler").font = Font(size=12, bold=True)
    ws.cell(row=sum_header_row, column=1).alignment = left

    # metrikler (2 sütun label-value x 4 satır)
    metrics = [
        ("Etüt Süresi (dk)", _num(etud_minutes,2)),
        ("Toplam Planlı Süre (dk)", _num(planned_sec/60,2)),
        ("Toplam Plansız Süre (dk)", _num(unplanned_sec/60,2)),
        ("Toplam Hatalı Süre (dk)", _num(faulty_sec/60,2)),
        ("Kapasite Kullanımı (%)", _num(utilization_pct,2)),
        ("Planlanan Üretim (adet)", _num(planned_qty,2)),
        ("Gerçekleşen Üretim (adet)", _num(actual_qty,2)),
        ("Gerçekleşme Oranı (%)", _num(performance_pct,2)),
    ]
    mrow = sum_header_row + 1
    idx = 0
    for rr in range(4):
        ws.cell(row=mrow+rr, column=1, value=metrics[idx][0]).font = bold
        ws.cell(row=mrow+rr, column=2, value=metrics[idx][1]); idx += 1
        ws.cell(row=mrow+rr, column=4, value=metrics[idx][0]).font = bold
        ws.cell(row=mrow+rr, column=5, value=metrics[idx][1]); idx += 1

    # duruş listesi başlık
    table_start = mrow + 5
    ws.merge_cells(start_row=table_start, start_column=1, end_row=table_start, end_column=7)
    ws.cell(row=table_start, column=1, value="Duruş Listesi").font = Font(size=12, bold=True)
    ws.cell(row=table_start, column=1).alignment = left

    headers = ["Duruş Türü", "Duruş Açıklaması", "Süre (sn)", "Süre (dk)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_start+1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # veri satırları
    data_start = table_start + 2
    error_data = list(error_data or [])
    if error_data:
        for i, e in enumerate(error_data):
            row_idx = data_start + i
            durus_turu = e.get("Duruş Türü", e.get("Hata Türü", ""))
            aciklama = e.get("Açıklama", "")
            sure_sn = e.get("Süre (sn)", 0) or 0
            ws.cell(row=row_idx, column=1, value=durus_turu).border = border
            ws.cell(row=row_idx, column=2, value=aciklama).border = border
            ws.cell(row=row_idx, column=3, value=sure_sn).border = border
            ws.cell(row=row_idx, column=4, value=round(float(sure_sn)/60, 2)).border = border
    else:
        row_idx = data_start
        ws.cell(row=row_idx, column=1, value="—").border = border
        ws.cell(row=row_idx, column=2, value="Duruş yok").border = border
        ws.cell(row=row_idx, column=3, value="").border = border
        ws.cell(row=row_idx, column=4, value="").border = border

    ws.freeze_panes = ws[f"A{data_start}"]

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path
